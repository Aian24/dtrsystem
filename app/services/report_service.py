"""
Report Generation Service — PDF (ReportLab) and Excel (OpenPyXL)
"""
from datetime import date
from pathlib import Path
from typing import Optional

from app.core.database import SessionLocal
from app.core.models import Employee, Company, AttendanceLog
from app.services.dtr_service import compute_dtr


# ─── PDF Generation ───────────────────────────────────────────────────────────

def generate_dtr_pdf(employee_id: int, date_from: date, date_to: date, output_path: str) -> None:
    """Generate a single employee DTR PDF."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    db = SessionLocal()
    try:
        emp = db.query(Employee).filter(Employee.id == employee_id).first()
        if not emp:
            raise ValueError(f"Employee {employee_id} not found")
        entries = compute_dtr(employee_id, date_from, date_to, db)
    finally:
        db.close()

    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    primary_blue = colors.HexColor("#2563EB")
    header_bg    = colors.HexColor("#F1F5F9")

    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontSize=14, textColor=primary_blue, spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#64748B"),
    )

    elements = []

    # Company / Employee header
    co_name = emp.company.name if emp.company else ""
    elements.append(Paragraph(co_name, title_style))
    elements.append(Paragraph("Daily Time Record (DTR)", sub_style))
    elements.append(Spacer(1, 10))

    info_data = [
        [
            Paragraph(f"<b>Employee:</b> {emp.last_name}, {emp.first_name}", styles["Normal"]),
            Paragraph(f"<b>Employee ID:</b> {emp.emp_id}", styles["Normal"]),
        ],
        [
            Paragraph(f"<b>Department:</b> {emp.department or '—'}", styles["Normal"]),
            Paragraph(f"<b>Period:</b> {date_from.strftime('%B %d, %Y')} – {date_to.strftime('%B %d, %Y')}", styles["Normal"]),
        ],
    ]
    info_table = Table(info_data, colWidths=["50%", "50%"])
    info_table.setStyle(TableStyle([
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 12))

    # DTR Table
    col_headers = ["Day", "Date", "Time In", "Break Out", "Break In", "Break Out 2", "Break In 2", "Time Out", "Late (min)", "Remarks"]
    table_data = [col_headers]

    for e in entries:
        d: date = e["date"]
        row = [
            d.strftime("%a"),
            d.strftime("%m/%d/%Y"),
            e.get("time_in")     or "",
            e.get("break_out_1") or "",
            e.get("break_in_1")  or "",
            e.get("break_out_2") or "",
            e.get("break_in_2")  or "",
            e.get("time_out")    or "",
            str(e.get("late_minutes") or ""),
            e.get("remarks")     or "",
        ]
        table_data.append(row)

    # Summaries
    table_data.append([""] * 10)
    present = sum(1 for e in entries if e.get("time_in"))
    absent  = sum(1 for e in entries if not e.get("time_in"))
    late    = sum(1 for e in entries if e.get("is_late"))
    table_data.append(["", "", "", "", "", "", "", f"Present: {present}", f"Absent: {absent}", f"Late: {late}"])

    col_widths = [1.4*cm, 2.2*cm, 2*cm, 2*cm, 2*cm, 2.2*cm, 2.2*cm, 2*cm, 2.2*cm, 2.4*cm]
    dtr_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    dtr_table.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0),  primary_blue),
        ("TEXTCOLOR",    (0,0), (-1,0),  colors.white),
        ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 8),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-3), [colors.white, header_bg]),
        ("GRID",         (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ("TOPPADDING",   (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0), (-1,-1), 4),
    ]))
    elements.append(dtr_table)

    # Signature line
    elements.append(Spacer(1, 20))
    sig_data = [
        ["_" * 35, "", "_" * 35],
        ["Employee Signature", "", "Verified by HR"],
    ]
    sig_table = Table(sig_data, colWidths=["45%", "10%", "45%"])
    sig_table.setStyle(TableStyle([
        ("ALIGN",   (0,0), (-1,-1), "CENTER"),
        ("FONTSIZE",(0,0), (-1,-1), 8),
        ("TEXTCOLOR",(0,1),(-1,1),  colors.HexColor("#64748B")),
    ]))
    elements.append(sig_table)

    doc.build(elements)


# ─── General Report Generation ────────────────────────────────────────────────

def generate_report(
    report_type: str,
    company_id: Optional[int],
    date_from: date,
    date_to: date,
    output_path: str,
    fmt: str = "pdf",
) -> None:
    """Route to the appropriate report generator."""
    db = SessionLocal()
    try:
        # Get employees
        q = db.query(Employee).filter(Employee.is_active == True)
        if company_id:
            q = q.filter(Employee.company_id == company_id)
        employees = q.all()

        rows = []
        for emp in employees:
            entries = compute_dtr(emp.id, date_from, date_to, db)
            present = sum(1 for e in entries if e.get("time_in"))
            absent  = sum(1 for e in entries if not e.get("time_in"))
            late    = sum(1 for e in entries if e.get("is_late"))
            rows.append({
                "emp_id":    emp.emp_id,
                "name":      f"{emp.last_name}, {emp.first_name}",
                "company":   emp.company.name if emp.company else "",
                "present":   present,
                "absent":    absent,
                "late":      late,
            })
    finally:
        db.close()

    if fmt == "pdf":
        _write_summary_pdf(rows, report_type, date_from, date_to, output_path)
    else:
        _write_summary_excel(rows, report_type, date_from, date_to, output_path)


def _write_summary_pdf(rows, report_type, date_from, date_to, output_path):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    styles = getSampleStyleSheet()
    primary = colors.HexColor("#2563EB")
    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    elements.append(Paragraph(f"{report_type.replace('_',' ').title()} Report", styles["Title"]))
    elements.append(Paragraph(f"{date_from} to {date_to}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    headers = ["Emp ID", "Name", "Company", "Present", "Absent", "Late"]
    data = [headers] + [
        [r["emp_id"], r["name"], r["company"], r["present"], r["absent"], r["late"]]
        for r in rows
    ]
    t = Table(data, colWidths=[2*cm, 5*cm, 4*cm, 2*cm, 2*cm, 2*cm], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), primary),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 8),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
    ]))
    elements.append(t)
    doc.build(elements)


def _write_summary_excel(rows, report_type, date_from, date_to, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = report_type.title()

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{report_type.replace('_',' ').title()} Report — {date_from} to {date_to}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="2563EB")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Emp ID", "Name", "Company", "Present", "Absent", "Late"]
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data
    for ri, row in enumerate(rows, 4):
        vals = [row["emp_id"], row["name"], row["company"], row["present"], row["absent"], row["late"]]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    # Column widths
    widths = [12, 28, 22, 10, 10, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)
